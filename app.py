import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from flask_sqlalchemy import SQLAlchemy
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from apscheduler.schedulers.background import BackgroundScheduler
import uuid
from werkzeug.utils import secure_filename
from io import BytesIO
from sqlalchemy import func, extract, case

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-key-change-this')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Email configuration (for reminders)
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', 'your-email@gmail.com')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', 'your-app-password')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_USERNAME', 'noreply@kenadmin.com')

mail = Mail(app)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)

# Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), default='staff')
    calendar_sync = db.Column(db.Boolean, default=False)
    reminder_preference = db.Column(db.String(20), default='email')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Budget(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    amount = db.Column(db.Float, nullable=False)
    department = db.Column(db.String(100))
    status = db.Column(db.String(20), default='pending')
    submitted_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    reviewed_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    review_comments = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Meeting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    date_time = db.Column(db.DateTime, nullable=False)
    duration = db.Column(db.Integer, default=60)
    meeting_link = db.Column(db.String(500))
    location = db.Column(db.String(200))
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    reminder_sent = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'))
    assigned_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    due_date = db.Column(db.DateTime, nullable=False)
    priority = db.Column(db.String(20), default='medium')
    status = db.Column(db.String(20), default='pending')
    meeting_id = db.Column(db.Integer, db.ForeignKey('meeting.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class MeetingAttendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    meeting_id = db.Column(db.Integer, db.ForeignKey('meeting.id'))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    status = db.Column(db.String(20), default='invited')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class CalendarEvent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    meeting_id = db.Column(db.Integer, db.ForeignKey('meeting.id'))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    event_id = db.Column(db.String(200))
    calendar_type = db.Column(db.String(50), default='ical')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Email reminder function
def send_meeting_reminder(meeting, user):
    try:
        msg = Message(
            f'Reminder: {meeting.title} - KEN Admin',
            recipients=[user.email]
        )
        msg.body = f"""
Hello {user.username},

This is a reminder for your upcoming meeting:

Meeting: {meeting.title}
Date: {meeting.date_time.strftime('%Y-%m-%d')}
Time: {meeting.date_time.strftime('%H:%M')}
Duration: {meeting.duration} minutes

Join link: {meeting.meeting_link or 'No link provided'}

You can view full details at: https://business-admin-1hpp.onrender.com/meeting/{meeting.id}

Best regards,
KEN Admin Team
"""
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False

# Task status update notification
def send_task_update_notification(task, user, old_status, new_status):
    try:
        msg = Message(
            f'Task Updated: {task.title} - KEN Admin',
            recipients=[user.email]
        )
        msg.body = f"""
Hello {user.username},

The status of your task has been updated.

Task: {task.title}
Old Status: {old_status}
New Status: {new_status}

View your tasks at: https://business-admin-1hpp.onrender.com/tasks

Best regards,
KEN Admin Team
"""
        mail.send(msg)
        return True
    except Exception as e:
        print(f"Failed to send update notification: {e}")
        return False

# Scheduled task to send reminders
scheduler = BackgroundScheduler()

def check_and_send_reminders():
    with app.app_context():
        now = datetime.utcnow()
        reminder_time = now + timedelta(hours=1)
        
        meetings = Meeting.query.filter(
            Meeting.date_time <= reminder_time,
            Meeting.date_time > now,
            Meeting.reminder_sent == False
        ).all()
        
        for meeting in meetings:
            attendees = MeetingAttendance.query.filter_by(meeting_id=meeting.id).all()
            for attendee in attendees:
                user = User.query.get(attendee.user_id)
                if user and user.reminder_preference == 'email':
                    send_meeting_reminder(meeting, user)
            meeting.reminder_sent = True
            db.session.commit()

# Start scheduler
scheduler.add_job(func=check_and_send_reminders, trigger="interval", minutes=30)
scheduler.start()


# Task assignment notification function
def send_task_assignment_notification(task, assigned_user, assigned_by_user):
    try:
        msg = Message(
            f'New Task Assigned: {task.title} - KEN Admin',
            recipients=[assigned_user.email]
        )
        msg.body = f"""
Hello {assigned_user.username},

A new task has been assigned to you by {assigned_by_user.username}.

Task Details:
━━━━━━━━━━━━━━━━━━━━━━
Title: {task.title}
Description: {task.description or 'No description provided'}
Priority: {task.priority.upper()}
Due Date: {task.due_date.strftime('%A, %B %d, %Y')}
Status: {task.status}

You can view and manage this task at:
https://business-admin-1hpp.onrender.com/tasks

━━━━━━━━━━━━━━━━━━━━━━
Please complete this task by the due date.

Best regards,
KEN Admin Team
"""
        mail.send(msg)
        print(f"✅ Task notification sent to {assigned_user.email}")
        return True
    except Exception as e:
        print(f"❌ Failed to send task notification: {e}")
        return False

# ============ ROUTES ============

@app.route('/')
@login_required
def dashboard():
    pending_budgets = Budget.query.filter_by(status='pending').count()
    approved_budgets = Budget.query.filter_by(status='approved').count()
    upcoming_meetings = Meeting.query.filter(Meeting.date_time > datetime.utcnow()).order_by(Meeting.date_time).limit(5).all()
    my_tasks = Task.query.filter_by(assigned_to=current_user.id, status='pending').order_by(Task.due_date).limit(10).all()
    recent_budgets = Budget.query.order_by(Budget.created_at.desc()).limit(5).all()
    return render_template('dashboard.html', 
                         pending_budgets=pending_budgets,
                         approved_budgets=approved_budgets,
                         upcoming_meetings=upcoming_meetings,
                         my_tasks=my_tasks,
                         recent_budgets=recent_budgets)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and check_password_hash(user.password, request.form['password']):
            login_user(user)
            flash(f'Welcome back, {user.username}!', 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid username or password', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        current_user.email = request.form.get('email', current_user.email)
        current_user.reminder_preference = request.form.get('reminder_preference', 'email')
        db.session.commit()
        flash('Profile updated!', 'success')
        return redirect(url_for('profile'))
    return render_template('profile.html')

@app.route('/calendar-sync')
@login_required
def calendar_sync():
    """Generate iCal file for user's meetings (only meetings user is invited to)"""
    # Get meetings where user is invited
    invited_meetings = Meeting.query.join(MeetingAttendance).filter(
        MeetingAttendance.user_id == current_user.id,
        Meeting.date_time > datetime.utcnow()
    ).order_by(Meeting.date_time).all()
    
    # Also include meetings created by user
    created_meetings = Meeting.query.filter(
        Meeting.created_by == current_user.id,
        Meeting.date_time > datetime.utcnow()
    ).all()
    
    # Combine and deduplicate
    all_meetings = list(set(invited_meetings + created_meetings))
    
    cal_data = """BEGIN:VCALENDAR
VERSION:2.0
PRODID:-//KEN Admin//Calendar//EN
CALSCALE:GREGORIAN
METHOD:PUBLISH
"""
    
    for meeting in all_meetings:
        dtstart = meeting.date_time.strftime('%Y%m%dT%H%M%SZ')
        dtend = (meeting.date_time + timedelta(minutes=meeting.duration)).strftime('%Y%m%dT%H%M%SZ')
        
        cal_data += f"""
BEGIN:VEVENT
UID:{meeting.id}@kenadmin.com
DTSTAMP:{datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')}
DTSTART:{dtstart}
DTEND:{dtend}
SUMMARY:{meeting.title}
DESCRIPTION:{meeting.description or ''}
LOCATION:{meeting.location or meeting.meeting_link or ''}
STATUS:CONFIRMED
SEQUENCE:0
END:VEVENT"""
    
    cal_data += "\nEND:VCALENDAR"
    
    return send_file(
        BytesIO(cal_data.encode('utf-8')),
        mimetype='text/calendar',
        as_attachment=True,
        download_name=f'KEN_Admin_Calendar_{datetime.utcnow().strftime("%Y%m%d")}.ics'
    )

# ============ MEETING ROUTES ============

@app.route('/meetings')
@login_required
def view_meetings():
    """Show only meetings that user has access to"""
    # Get meetings where user is invited
    invited_meetings = Meeting.query.join(MeetingAttendance).filter(
        MeetingAttendance.user_id == current_user.id
    ).order_by(Meeting.date_time.desc()).all()
    
    # Get meetings created by user
    created_meetings = Meeting.query.filter_by(
        created_by=current_user.id
    ).order_by(Meeting.date_time.desc()).all()
    
    # Combine and deduplicate
    all_meetings = list(set(invited_meetings + created_meetings))
    
    # If admin, show all meetings
    if current_user.role == 'admin':
        all_meetings = Meeting.query.order_by(Meeting.date_time.desc()).all()
    
    return render_template('meetings.html', meetings=all_meetings)

@app.route('/meeting/create', methods=['GET', 'POST'])
@login_required
def create_meeting():
    if request.method == 'POST':
        meeting = Meeting(
            title=request.form['title'],
            description=request.form.get('description', ''),
            date_time=datetime.strptime(request.form['date_time'], '%Y-%m-%dT%H:%M'),
            duration=int(request.form.get('duration', 60)),
            meeting_link=request.form.get('meeting_link', ''),
            location=request.form.get('location', ''),
            created_by=current_user.id
        )
        db.session.add(meeting)
        db.session.commit()
        
        # Add creator as attendee
        attendance = MeetingAttendance(meeting_id=meeting.id, user_id=current_user.id, status='confirmed')
        db.session.add(attendance)
        
        # Add selected invitees
        invitees = request.form.getlist('invitees')
        for user_id in invitees:
            if int(user_id) != current_user.id:
                attendance = MeetingAttendance(meeting_id=meeting.id, user_id=int(user_id), status='invited')
                db.session.add(attendance)
        
        db.session.commit()
        flash(f'Meeting scheduled successfully! {len(invitees)} people invited.', 'success')
        return redirect(url_for('view_meetings'))
    
    users = User.query.all()
    return render_template('create_meeting.html', users=users)

@app.route('/meeting/<int:id>')
@login_required
def view_meeting(id):
    meeting = Meeting.query.get_or_404(id)
    
    # Check access
    attendance = MeetingAttendance.query.filter_by(
        meeting_id=meeting.id, 
        user_id=current_user.id
    ).first()
    
    has_access = (attendance is not None) or (meeting.created_by == current_user.id) or (current_user.role == 'admin')
    
    if not has_access:
        return render_template('view_meeting.html', meeting=meeting, has_access=False)
    
    # Get attendees
    attendees = db.session.query(
        User.id, User.username, User.email, User.role,
        MeetingAttendance.status
    ).join(MeetingAttendance, User.id == MeetingAttendance.user_id)\
     .filter(MeetingAttendance.meeting_id == meeting.id).all()
    
    creator = User.query.get(meeting.created_by)
    creator_name = creator.username if creator else 'Unknown'
    all_users = User.query.all()
    attendee_ids = [att.id for att in attendees]
    
    return render_template('view_meeting.html', 
                         meeting=meeting, 
                         attendees=attendees,
                         creator_name=creator_name,
                         all_users=all_users,
                         attendee_ids=attendee_ids,
                         has_access=True)

@app.route('/meeting/<int:id>/invite', methods=['POST'])
@login_required
def invite_to_meeting(id):
    """Add more attendees to an existing meeting"""
    meeting = Meeting.query.get_or_404(id)
    
    if meeting.created_by != current_user.id and current_user.role != 'admin':
        flash('Only the meeting organizer can invite attendees.', 'danger')
        return redirect(url_for('view_meeting', id=id))
    
    user_ids = request.form.getlist('user_ids')
    invited_count = 0
    
    for user_id in user_ids:
        existing = MeetingAttendance.query.filter_by(meeting_id=id, user_id=int(user_id)).first()
        if not existing:
            attendance = MeetingAttendance(meeting_id=id, user_id=int(user_id), status='invited')
            db.session.add(attendance)
            invited_count += 1
    
    db.session.commit()
    flash(f'{invited_count} new invitation(s) sent!', 'success')
    return redirect(url_for('view_meeting', id=id))

@app.route('/meeting/<int:id>/ical')
@login_required
def download_ical(id):
    """Download single meeting as iCal file"""
    meeting = Meeting.query.get_or_404(id)
    
    attendance = MeetingAttendance.query.filter_by(meeting_id=meeting.id, user_id=current_user.id).first()
    if not attendance and meeting.created_by != current_user.id and current_user.role != 'admin':
        flash('You are not invited to this meeting.', 'danger')
        return redirect(url_for('view_meetings'))
    
    dtstart = meeting.date_time.strftime('%Y%m%dT%H%M%SZ')
    dtend = (meeting.date_time + timedelta(minutes=meeting.duration)).strftime('%Y%m%dT%H%M%SZ')
    
    cal_data = f"""BEGIN:VCALENDAR
VERSION:2.0
PRODID:-//KEN Admin//Meeting//EN
CALSCALE:GREGORIAN
BEGIN:VEVENT
UID:{meeting.id}@kenadmin.com
DTSTAMP:{datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')}
DTSTART:{dtstart}
DTEND:{dtend}
SUMMARY:{meeting.title}
DESCRIPTION:{meeting.description or ''}
LOCATION:{meeting.location or meeting.meeting_link or ''}
STATUS:CONFIRMED
SEQUENCE:0
BEGIN:VALARM
TRIGGER:-PT1H
ACTION:DISPLAY
DESCRIPTION:Reminder for {meeting.title}
END:VALARM
END:VEVENT
END:VCALENDAR"""
    
    return send_file(
        BytesIO(cal_data.encode('utf-8')),
        mimetype='text/calendar',
        as_attachment=True,
        download_name=f'{meeting.title.replace(" ", "_")}.ics'
    )

@app.route('/meeting/<int:id>/delete', methods=['POST'])
@login_required
def delete_meeting(id):
    """Delete a meeting (creator or admin only)"""
    meeting = Meeting.query.get_or_404(id)
    
    if meeting.created_by != current_user.id and current_user.role != 'admin':
        flash('You do not have permission to delete this meeting.', 'danger')
        return redirect(url_for('view_meetings'))
    
    try:
        meeting_title = meeting.title
        
        MeetingAttendance.query.filter_by(meeting_id=meeting.id).delete()
        CalendarEvent.query.filter_by(meeting_id=meeting.id).delete()
        Task.query.filter_by(meeting_id=meeting.id).update({'meeting_id': None})
        
        db.session.delete(meeting)
        db.session.commit()
        
        flash(f'Meeting "{meeting_title}" has been deleted successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting meeting: {str(e)}', 'danger')
    
    return redirect(url_for('view_meetings'))

@app.route('/budgets')
@login_required
def view_budgets():
    if current_user.role == 'admin':
        budgets = Budget.query.order_by(Budget.created_at.desc()).all()
    else:
        budgets = Budget.query.filter_by(submitted_by=current_user.id).order_by(Budget.created_at.desc()).all()
    return render_template('budgets.html', budgets=budgets)

@app.route('/budget/create', methods=['GET', 'POST'])
@login_required
def create_budget():
    if request.method == 'POST':
        budget = Budget(
            title=request.form['title'],
            description=request.form.get('description', ''),
            amount=float(request.form['amount']),
            department=request.form.get('department', ''),
            submitted_by=current_user.id
        )
        db.session.add(budget)
        db.session.commit()
        flash('Budget submitted successfully!', 'success')
        return redirect(url_for('view_budgets'))
    return render_template('create_budget.html')

@app.route('/budget/<int:id>')
@login_required
def view_budget(id):
    budget = Budget.query.get_or_404(id)
    return render_template('view_budget.html', budget=budget)

@app.route('/budget/<int:id>/review', methods=['POST'])
@login_required
def review_budget(id):
    if current_user.role != 'admin':
        flash('Only admins can review budgets.', 'danger')
        return redirect(url_for('view_budgets'))
    budget = Budget.query.get_or_404(id)
    action = request.form.get('action')
    budget.review_comments = request.form.get('comments', '')
    budget.reviewed_by = current_user.id
    if action == 'approve':
        budget.status = 'approved'
        flash(f'Budget "{budget.title}" approved!', 'success')
    elif action == 'reject':
        budget.status = 'rejected'
        flash(f'Budget "{budget.title}" rejected.', 'warning')
    db.session.commit()
    return redirect(url_for('view_budgets'))

@app.route('/tasks')
@login_required
def view_tasks():
    if current_user.role == 'admin':
        tasks = Task.query.order_by(Task.due_date).all()
    else:
        tasks = Task.query.filter_by(assigned_to=current_user.id).order_by(Task.due_date).all()
    return render_template('tasks.html', tasks=tasks)

@app.route('/task/create', methods=['GET', 'POST'])
@login_required
def create_task():
    if request.method == 'POST':
        task = Task(
            title=request.form['title'],
            description=request.form.get('description', ''),
            assigned_to=int(request.form['assigned_to']),
            assigned_by=current_user.id,
            due_date=datetime.strptime(request.form['due_date'], '%Y-%m-%d'),
            priority=request.form.get('priority', 'medium')
        )
        db.session.add(task)
        db.session.commit()
        flash('Task assigned successfully!', 'success')
        return redirect(url_for('view_tasks'))
    users = User.query.filter(User.id != current_user.id).all()
    return render_template('create_task.html', users=users)

@app.route('/task/<int:id>/update', methods=['POST'])
@login_required
def update_task(id):
    task = Task.query.get_or_404(id)
    if task.assigned_to == current_user.id or current_user.role == 'admin':
        old_status = task.status
        new_status = request.form.get('status')
        
        if old_status != new_status:
            task.status = new_status
            db.session.commit()
            
            # Notify the assigned user about status change
            assigned_user = User.query.get(task.assigned_to)
            if assigned_user and assigned_user.reminder_preference == 'email' and assigned_user.id != current_user.id:
                send_task_update_notification(task, assigned_user, old_status, new_status)
            
            flash(f'Task updated from "{old_status}" to "{new_status}"!', 'success')
        else:
            flash('Task updated!', 'success')
    
    return redirect(url_for('view_tasks'))

@app.route('/api/check-new-tasks')
@login_required
def check_new_tasks():
    """API endpoint for checking new tasks (for real-time notifications)"""
    # Get tasks assigned in the last 5 minutes
    five_minutes_ago = datetime.utcnow() - timedelta(minutes=5)
    new_tasks = Task.query.filter(
        Task.assigned_to == current_user.id,
        Task.created_at >= five_minutes_ago,
        Task.status == 'pending'
    ).count()
    
    return jsonify({'new_tasks': new_tasks})

@app.route('/reports')
@login_required
def reports():
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))
    
    # ============ BUDGET REPORTS ============
    total_budgets = Budget.query.count()
    pending_budgets = Budget.query.filter_by(status='pending').count()
    approved_budgets_count = Budget.query.filter_by(status='approved').count()
    rejected_budgets_count = Budget.query.filter_by(status='rejected').count()
    
    # Department-wise budget summary
    dept_budgets = db.session.query(
        Budget.department,
        func.count(Budget.id).label('count'),
        func.sum(Budget.amount).label('total_amount'),
        func.sum(case((Budget.status == 'approved', Budget.amount), else_=0)).label('approved_amount'),
        func.sum(case((Budget.status == 'pending', Budget.amount), else_=0)).label('pending_amount')
    ).group_by(Budget.department).all()
    
    # Total approved amount across all departments
    approved_budgets = Budget.query.filter_by(status='approved').all()
    total_approved_amount = sum(b.amount for b in approved_budgets)
    
    # ============ TASK REPORTS ============
    total_tasks = Task.query.count()
    completed_tasks = Task.query.filter_by(status='completed').count()
    pending_tasks = Task.query.filter_by(status='pending').count()
    in_progress_tasks = Task.query.filter_by(status='in_progress').count()
    
    # Task completion rate
    task_completion_rate = (completed_tasks / total_tasks * 100) if total_tasks > 0 else 0
    
    # Tasks by priority
    high_priority_tasks = Task.query.filter_by(priority='high', status='pending').count()
    medium_priority_tasks = Task.query.filter_by(priority='medium', status='pending').count()
    low_priority_tasks = Task.query.filter_by(priority='low', status='pending').count()
    
    # Tasks by assignee
    tasks_by_user = db.session.query(
        User.username,
        func.count(Task.id).label('task_count'),
        func.sum(case((Task.status == 'completed', 1), else_=0)).label('completed_count')
    ).outerjoin(Task, User.id == Task.assigned_to)\
     .group_by(User.id, User.username).all()
    
    # ============ MEETING REPORTS ============
    total_meetings = Meeting.query.count()
    upcoming_meetings = Meeting.query.filter(Meeting.date_time > datetime.utcnow()).count()
    past_meetings = Meeting.query.filter(Meeting.date_time <= datetime.utcnow()).count()
    
    # Meetings by month (last 6 months)
    meetings_by_month = db.session.query(
        extract('year', Meeting.date_time).label('year'),
        extract('month', Meeting.date_time).label('month'),
        func.count(Meeting.id).label('count')
    ).group_by('year', 'month').order_by('year', 'month').limit(6).all()
    
    # ============ ACTIVITY SUMMARY ============
    # Recent activities (last 30 days)
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    
    recent_budgets = Budget.query.filter(Budget.created_at >= thirty_days_ago).count()
    recent_meetings = Meeting.query.filter(Meeting.created_at >= thirty_days_ago).count()
    recent_tasks = Task.query.filter(Task.created_at >= thirty_days_ago).count()
    completed_recent_tasks = Task.query.filter(
        Task.status == 'completed',
        Task.created_at >= thirty_days_ago
    ).count()
    
    # Top performing departments (by approved budget amount)
    top_departments = db.session.query(
        Budget.department,
        func.sum(Budget.amount).label('total')
    ).filter(Budget.status == 'approved')\
     .group_by(Budget.department)\
     .order_by(func.sum(Budget.amount).desc()).limit(5).all()
    
    return render_template('reports.html', 
                         # Budget stats
                         total_budgets=total_budgets,
                         pending_budgets=pending_budgets,
                         approved_budgets_count=approved_budgets_count,
                         rejected_budgets_count=rejected_budgets_count,
                         total_approved_amount=total_approved_amount,
                         dept_budgets=dept_budgets,
                         top_departments=top_departments,
                         
                         # Task stats
                         total_tasks=total_tasks,
                         completed_tasks=completed_tasks,
                         pending_tasks=pending_tasks,
                         in_progress_tasks=in_progress_tasks,
                         task_completion_rate=task_completion_rate,
                         high_priority_tasks=high_priority_tasks,
                         medium_priority_tasks=medium_priority_tasks,
                         low_priority_tasks=low_priority_tasks,
                         tasks_by_user=tasks_by_user,
                         
                         # Meeting stats
                         total_meetings=total_meetings,
                         upcoming_meetings=upcoming_meetings,
                         past_meetings=past_meetings,
                         meetings_by_month=meetings_by_month,
                         
                         # Activity stats
                         recent_budgets=recent_budgets,
                         recent_meetings=recent_meetings,
                         recent_tasks=recent_tasks,
                         completed_recent_tasks=completed_recent_tasks)

# ============ EXPORT ROUTES ============

@app.route('/reports/export/pdf')
@login_required
def export_pdf():
    """Export reports as PDF"""
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))
    
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from io import BytesIO
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=30)
    story.append(Paragraph("KEN Admin - Business Report", title_style))
    story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Budgets Summary
    story.append(Paragraph("Budgets Summary", styles['Heading2']))
    budgets = Budget.query.all()
    budget_data = [['Department', 'Title', 'Amount', 'Status', 'Date']]
    for b in budgets:
        budget_data.append([
            b.department or 'N/A',
            b.title[:30],
            f'${b.amount:,.2f}',
            b.status,
            b.created_at.strftime('%Y-%m-%d')
        ])
    
    budget_table = Table(budget_data)
    budget_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    story.append(budget_table)
    story.append(Spacer(1, 20))
    
    # Tasks Summary
    story.append(Paragraph("Tasks Summary", styles['Heading2']))
    tasks = Task.query.all()
    task_data = [['Title', 'Priority', 'Status', 'Due Date']]
    for t in tasks:
        task_data.append([t.title[:30], t.priority, t.status, t.due_date.strftime('%Y-%m-%d')])
    
    task_table = Table(task_data)
    task_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    story.append(task_table)
    
    doc.build(story)
    buffer.seek(0)
    
    return send_file(
        buffer, 
        as_attachment=True, 
        download_name=f'KEN_Admin_Report_{datetime.now().strftime("%Y%m%d")}.pdf', 
        mimetype='application/pdf'
    )


@app.route('/reports/export/excel')
@login_required
def export_excel():
    """Export reports as Excel"""
    if current_user.role != 'admin':
        flash('Access denied', 'danger')
        return redirect(url_for('dashboard'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    wb = Workbook()
    
    # Budgets sheet
    ws_budgets = wb.active
    ws_budgets.title = "Budgets"
    
    # Headers
    headers = ['ID', 'Department', 'Title', 'Amount', 'Status', 'Submitted By', 'Created At']
    for col, header in enumerate(headers, 1):
        cell = ws_budgets.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="6b46c0", end_color="6b46c0", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    for row, budget in enumerate(Budget.query.all(), 2):
        ws_budgets.cell(row=row, column=1, value=budget.id)
        ws_budgets.cell(row=row, column=2, value=budget.department or 'N/A')
        ws_budgets.cell(row=row, column=3, value=budget.title)
        ws_budgets.cell(row=row, column=4, value=budget.amount)
        ws_budgets.cell(row=row, column=5, value=budget.status)
        ws_budgets.cell(row=row, column=6, value=budget.submitted_by)
        ws_budgets.cell(row=row, column=7, value=budget.created_at.strftime('%Y-%m-%d %H:%M'))
    
    # Adjust column widths
    for column in ws_budgets.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_budgets.column_dimensions[column_letter].width = adjusted_width
    
    # Tasks sheet
    ws_tasks = wb.create_sheet("Tasks")
    task_headers = ['ID', 'Title', 'Priority', 'Status', 'Assigned To', 'Due Date']
    for col, header in enumerate(task_headers, 1):
        cell = ws_tasks.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="6b46c0", end_color="6b46c0", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, task in enumerate(Task.query.all(), 2):
        ws_tasks.cell(row=row, column=1, value=task.id)
        ws_tasks.cell(row=row, column=2, value=task.title)
        ws_tasks.cell(row=row, column=3, value=task.priority)
        ws_tasks.cell(row=row, column=4, value=task.status)
        ws_tasks.cell(row=row, column=5, value=task.assigned_to)
        ws_tasks.cell(row=row, column=6, value=task.due_date.strftime('%Y-%m-%d'))
    
    # Meetings sheet
    ws_meetings = wb.create_sheet("Meetings")
    meeting_headers = ['ID', 'Title', 'Date & Time', 'Duration', 'Meeting Link', 'Location']
    for col, header in enumerate(meeting_headers, 1):
        cell = ws_meetings.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="6b46c0", end_color="6b46c0", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for row, meeting in enumerate(Meeting.query.all(), 2):
        ws_meetings.cell(row=row, column=1, value=meeting.id)
        ws_meetings.cell(row=row, column=2, value=meeting.title)
        ws_meetings.cell(row=row, column=3, value=meeting.date_time.strftime('%Y-%m-%d %H:%M'))
        ws_meetings.cell(row=row, column=4, value=meeting.duration)
        ws_meetings.cell(row=row, column=5, value=meeting.meeting_link or 'N/A')
        ws_meetings.cell(row=row, column=6, value=meeting.location or 'N/A')
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer, 
        as_attachment=True, 
        download_name=f'KEN_Admin_Report_{datetime.now().strftime("%Y%m%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ============ USER MANAGEMENT ROUTES ============

@app.route('/users')
@login_required
def view_users():
    """View all users (admin only)"""
    if current_user.role != 'admin':
        flash('Access denied. Only admins can view users.', 'danger')
        return redirect(url_for('dashboard'))
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('users.html', users=users)

@app.route('/user/create', methods=['POST'])
@login_required
def create_user():
    """Create a new user (admin only)"""
    if current_user.role != 'admin':
        flash('Access denied. Only admins can create users.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get form data
    username = request.form.get('username', '').strip()
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '')
    role = request.form.get('role', 'staff')
    
    # Validation
    if not username or not email or not password:
        flash('All fields are required.', 'danger')
        return redirect(url_for('view_users'))
    
    # Check if username already exists
    existing_user = User.query.filter_by(username=username).first()
    if existing_user:
        flash(f'Username "{username}" already exists. Please choose a different username.', 'danger')
        return redirect(url_for('view_users'))
    
    # Check if email already exists
    existing_email = User.query.filter_by(email=email).first()
    if existing_email:
        flash(f'Email "{email}" already exists. Please use a different email.', 'danger')
        return redirect(url_for('view_users'))
    
    # Password length validation
    if len(password) < 4:
        flash('Password must be at least 4 characters long.', 'danger')
        return redirect(url_for('view_users'))
    
    # Create new user
    try:
        user = User(
            username=username,
            email=email,
            password=generate_password_hash(password),
            role=role,
            reminder_preference='email'
        )
        
        db.session.add(user)
        db.session.commit()
        
        flash(f'User "{user.username}" has been created successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating user: {str(e)}', 'danger')
    
    return redirect(url_for('view_users'))

@app.route('/user/delete/<int:id>', methods=['POST'])
@login_required
def delete_user(id):
    """Delete a user (admin only)"""
    if current_user.role != 'admin':
        flash('Access denied. Only admins can delete users.', 'danger')
        return redirect(url_for('view_users'))
    
    user_to_delete = User.query.get_or_404(id)
    
    # Prevent admin from deleting themselves
    if user_to_delete.id == current_user.id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('view_users'))
    
    # Prevent deleting the default admin account
    if user_to_delete.username == 'admin':
        flash('Cannot delete the default admin account.', 'danger')
        return redirect(url_for('view_users'))
    
    try:
        username = user_to_delete.username
        
        # Handle related records before deletion
        # 1. Update budgets submitted by this user
        Budget.query.filter_by(submitted_by=user_to_delete.id).update({'submitted_by': None})
        
        # 2. Update meetings created by this user
        Meeting.query.filter_by(created_by=user_to_delete.id).update({'created_by': None})
        
        # 3. Update tasks assigned to this user
        Task.query.filter_by(assigned_to=user_to_delete.id).update({'assigned_to': None})
        
        # 4. Delete meeting attendance records
        MeetingAttendance.query.filter_by(user_id=user_to_delete.id).delete()
        
        # 5. Delete calendar events
        CalendarEvent.query.filter_by(user_id=user_to_delete.id).delete()
        
        # Finally delete the user
        db.session.delete(user_to_delete)
        db.session.commit()
        
        flash(f'User "{username}" has been deleted successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting user: {str(e)}', 'danger')
    
    return redirect(url_for('view_users'))

# Initialize database
with app.app_context():
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', email='admin@business.com', 
                    password=generate_password_hash('admin123'), role='admin', reminder_preference='email')
        db.session.add(admin)
        staff = User(username='staff', email='staff@business.com', 
                    password=generate_password_hash('staff123'), role='staff', reminder_preference='email')
        db.session.add(staff)
        db.session.commit()
        print("✅ KEN Admin initialized with default users")

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)